# --- remove the HASH and save as requirements.txt in the same directory

# openpyxl
# pandas
# PyPDF2
# colorama
# Pillow
# reportlab
# pymupdf


# --- remove the HASH and save as requirements.txt in the same directory

# To install all requirement library run it in console by removing HASH:
# pip install -r requirements.txt --trusted-host pypi.org --trusted-host pypi.python.org --trusted-host=files.pythonhosted.org


import csv
import os
import time
from datetime import datetime

import fitz
import openpyxl
import pandas as pd
import PyPDF2
from colorama import Back, Style
from openpyxl import load_workbook
from PIL import Image
from PyPDF2 import PdfMerger
from reportlab.lib.pagesizes import A4, letter
from reportlab.pdfgen import canvas


def landing():
    os.system('cls' if os.name == 'nt' else 'clear')
    print("\nPython for handling Excel and PDF files")
    print(Back.GREEN+"Imtiaz Rashid initiated this project, and Saklayen Ahmed took it to the next level."+Style.RESET_ALL)
    print("\nPlease wait. Loading...")
    time.sleep(3)
    os.system('cls' if os.name == 'nt' else 'clear')
    print("-----------------------------------")
    print("Welcome to Python for Excel and PDF")
    print("-----------------------------------")

def initial():
    landing()
    intinput = input(
        '''\nWhat do you want:\n\n'''
        '''1. [EXCEL] Merge excel files\n'''
        '''2. [EXCEL] Clean excel files\n'''
        '''3. [EXCEL] Clean and Merge excel files\n'''
        '''4. [EXCEL] Filter with keyword and Merge excel files\n'''
        '''5. [EXCEL] Merge excel sheets\n'''
        '''6. [PDF] Multiple PDF Merge\n'''
        '''7. [PDF] Split PDF Defined pages\n'''
        '''8. [PDF] Split PDF all pages\n'''
        '''9. [JPG] Convert JPG to PDF\n'''
        '''10.[JPG] Convert PDF to JPG\n'''
        '''00. Report an Issue\n'''
        '''\nChoose your option (1/2/3/4/5/6/7/8/9/10):\n\n'''
        )
    
    os.system('cls' if os.name == 'nt' else 'clear')

    if intinput=="1":
        intext = "\nMerge Excel files\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        mergeExcel()
    elif intinput =="2":
        intext = "\nClean Excel files\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        cleanExcel()
    elif intinput =="3":
        intext = "\nClean and Merge Excel files"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        clean_and_merge()
    elif intinput =="4":
        intext = "\nFilter and Merge Excel files\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        filter_and_merge_key()
    elif intinput =="5":
        intext = "\nMerge Excel Sheets\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        merge_sheets()
    elif intinput =="6":
        intext = "\nMultiple PDF Merge\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        mergePDF()
    elif intinput =="7":
        intext = "\nSplit PDF Defined pages\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        pdfSplitDefined()
    elif intinput =="8":
        intext = "\nSplit PDF all pages\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        pdfSplitAll()
    elif intinput =="9":
        intext = "\nConvert JPG to PDF\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        jpgtopdf()
    elif intinput =="10":
        intext = "\nConvert PDF to JPG\n"
        intext = Back.GREEN + intext + Style.RESET_ALL
        print(intext)
        pdftojpg()
    elif intinput =="00":
        print(Back.RED + "\nReport an Issue" + Style.RESET_ALL)
        print("\nPlease feel free to reach out if you have any feedback on this Python script.\n")
        print("You can reach me either by direct message or by creating an issue on GitHub.")
        url = "https://github.com/saklayenahmed/python-with-excel-pdf/issues/new"
        intext = "\nGithub Issue: Press and hold 'alt' then click on this link. Link: "+ Back.GREEN + url + Style.RESET_ALL
        print(intext)
    elif not intinput.isdigit() or not (1 <= int(intinput) <= 8):
        if not intinput.isdigit():
            intext = "\nSorry! You have to select any option here. Executing again..."
        else:
            intext = "\nTo select any option, you have to enter a number between 1 and 8. Executing again..."
        intext = Back.RED + intext + Style.RESET_ALL
        print(intext)
        initial()
    
    continue_input = input("\n\n"+ Back.GREEN + "Press Enter to continue. " + Style.RESET_ALL)

    if not continue_input:
        initial()


def pdftojpg():
    RESOLUTION = 700
    curr_time = datetime.now().strftime("%Y%m%d%H%M%S")
    input_folder = input("\nEnter PDF folder directory: ")
    pdf_files = [os.path.join(input_folder, file) for file in os.listdir(input_folder) if file.lower().endswith(('.pdf', '.PDF'))]
    pdf_files.sort()

    if not pdf_files:
        print(f"\n No PDF files found in this folder.")
        return
    else:
        print("\nPDF files found in the folder:\n")
        for index, pdf_file in enumerate(pdf_files, start=1):
            print(f"{index}. {os.path.basename(pdf_file)}")


    converted = False

    while not converted:
        pdf_choice = input("\nEnter PDF folder directory or type 'cancel': ")

        if pdf_choice=='cancel' or pdf_choice =="":
            print("PDF to JPG conversion canceled, thank you !")
            return

        pdf_index = int(pdf_choice) -1
        if pdf_index < 0 and pdf_index >= len(pdf_files):
            print("Invalid choice.")
            return

        selected_pdf = pdf_files[pdf_index]
        input_pdf_directory = os.path.join(input_folder,selected_pdf)

        pdf = fitz.open(input_pdf_directory)

        root, extension = os.path.splitext(input_pdf_directory)

        dir_name = f"{input_folder}/{os.path.basename(root)}_{curr_time}"
        os.makedirs(f"{dir_name}")
        
        print(f"\n PDF converting to JPG. Please wait a moment...!\n")

        for page_no in range(pdf.page_count):
            page = pdf[page_no]
            image = page.get_pixmap(matrix=fitz.Matrix(RESOLUTION / 72, RESOLUTION / 72))
            img_path = f"{dir_name}/page_{page_no+1}.jpg"
            image.save(img_path)
        print(f"\n({os.path.basename(selected_pdf)}) Converted successfully and JPG saved to : {dir_name}")
        pdf.close()
        converted=False


def jpgtopdf():
    curr_time = datetime.now().strftime("%Y%m%d%H%M%S")
    converted = False
    while not converted:
        input_folder = input("\nEnter Image folder directory or type 'cancel':\n")
        if input_folder=='cancel' or input_folder=="":
            print("Conversion canceled, thank you !")
            return
        image_paths = [os.path.join(input_folder, file) for file in os.listdir(input_folder) if file.lower().endswith(('.jpg', '.png', '.jpeg', '.gif', '.bmp'))]
        pdf_output = os.path.join(input_folder, f"{os.path.basename(input_folder)}_"+curr_time+".pdf")
        c = canvas.Canvas(pdf_output, pagesize=letter)
        
        print(f"\n JPG converting to PDF. Please wait a moment...!\n")

        for image_path in image_paths:
            image = Image.open(image_path)
            image_width, image_height = image.size

            if image_width < A4[0] and image_height < A4[1]:
                page_size = A4
            else:
                page_size = (image_width, image_height)

            c.setPageSize(page_size)
            c.drawImage(image_path, 0, 0, width=image_width, height=image_height, preserveAspectRatio=True, anchor='nw')
            c.showPage()

        c.save()

    print(f"\nPDF created and save to : {pdf_output}")
    converted = False


def pdfSplitAll():
    curr_time = datetime.now().strftime("%Y%m%d%H%M%S")
    
    def split_pdf(input_path, output_folder, pages):
        pdf = PyPDF2.PdfReader(input_path)
        
        for page_num in range(len(pdf.pages)):
            output_pdf_name = f"Page_{page_num + 1}_{curr_time}.pdf"
            output_pdf_path = os.path.join(output_folder, output_pdf_name)
            
            with open(output_pdf_path, "wb") as output_file:
                writer = PyPDF2.PdfWriter()
                writer.add_page(pdf.pages[page_num])
                writer.write(output_file)

    folder_path = input("\nEnter PDF folder directory:\n")
    pdf_files = [pdf_file for pdf_file in os.listdir(folder_path) if pdf_file.lower().endswith('.pdf')]
    pdf_files.sort()

    if not pdf_files:
        print("No PDF files found in the folder.")
        return

    print("\nPDF files found in the folder:\n")
    for index, pdf_file in enumerate(pdf_files, start=1):
        print(f"{index}. {pdf_file}")

    pdf_choice = input("\nEnter the number of the PDF to split or type 'cancel': ")
    if pdf_choice.lower() == "cancel":
        print("\nPDF splitting canceled. Thank you!")
        return

    try:
        pdf_index = int(pdf_choice) - 1
        if pdf_index < 0 or pdf_index >= len(pdf_files):
            print("Invalid PDF number.")
            return

        selected_pdf = pdf_files[pdf_index]
        input_pdf_path = os.path.join(folder_path, selected_pdf)
        output_folder = os.path.join(folder_path, f"{os.path.splitext(selected_pdf)[0]}_{curr_time}")
        os.makedirs(output_folder, exist_ok=True)
        
        pages_to_split = input("\nDo you want to split these PDFs? (yes/no): ")

        print(f"\n PDF Splitting have been started. Please wait a moment...!\n")

        if pages_to_split.lower() == "yes":
            split_pdf(input_pdf_path, output_folder, pages_to_split)
            print(f"\nPDF split into individual pages and saved to {output_folder}")

    except ValueError:
        print("Invalid input.")

def pdfSplitDefined():
    curr_time = datetime.now().strftime("%Y%m%d%H%M%S")
    
    def split_pdf(input_path, output_path, pages):
        pdf = PyPDF2.PdfReader(input_path)
        
        with open(output_path, "wb") as output_file:
            writer = PyPDF2.PdfWriter()
            
            for page_range in pages.split(','):
                if '-' in page_range:
                    start, end = map(int, page_range.split('-'))
                    for page_num in range(start - 1, end):
                        writer.add_page(pdf.pages[page_num])
                else:
                    page_num = int(page_range) - 1
                    writer.add_page(pdf.pages[page_num])
            
            writer.write(output_file)

    folder_path = input("\nEnter PDF folder directory:\n")
    pdf_files = [pdf_file for pdf_file in os.listdir(folder_path) if pdf_file.endswith('.pdf')]
    pdf_files.sort()

    if not pdf_files:
        print("No PDF files found in the folder.")
        return

    print("\nPDF files found in the folder:\n")
    for index, pdf_file in enumerate(pdf_files, start=1):
        print(f"{index}. {pdf_file}")

    pdf_choice = input("\nEnter the number of the PDF to split or type 'cancel': ")
    if pdf_choice.lower() == "cancel":
        print("\nPDF splitting canceled. Thank you!")
        return

    try:
        pdf_index = int(pdf_choice) - 1
        if pdf_index < 0 or pdf_index >= len(pdf_files):
            print("Invalid PDF number.")
            return

        selected_pdf = pdf_files[pdf_index]
        input_pdf_path = os.path.join(folder_path, selected_pdf)

        pages_to_split = input("\nEnter page numbers or ranges (e.g., 1-3, 5, 7-9):\n")

        # Generate the output PDF path based on the input PDF file name
        output_pdf_name = f"Split_{os.path.splitext(selected_pdf)[0]}_{curr_time}.pdf"
        output_pdf_path = os.path.join(folder_path, output_pdf_name)
        
        print(f"\n PDF Splitting have been started. Please wait a moment...!\n")

        split_pdf(input_pdf_path, output_pdf_path, pages_to_split)
        print(f"\nPDF pages split and saved to {output_pdf_path}")

    except ValueError:
        print("Invalid input.")

def mergePDF():
    curr_time = datetime.now().strftime("%Y%m%d%H%M%S")
    pdf_directory = input("Enter PDF folder directory:\n")
    folder_name = os.path.basename(pdf_directory)
    pdf_files = [pdf_file for pdf_file in os.listdir(pdf_directory) if pdf_file.lower().endswith('.pdf')]
    pdf_files.sort()

    if not pdf_files:
        print("\nNo PDF files found in the directory.")
        return

    print("\nPDF files found in the directory:\n")
    for pdf_file in pdf_files:
        print(pdf_file)

    confirmation = input("\nDo you want to merge these PDFs? (yes/no): ")
    if confirmation.lower() != "yes" or confirmation=="":
        print("\nPDF merging canceled. Thank you!")
        return

    pdf_merger = PdfMerger()

    print(f"\n PDF Merging have been started. Please wait a moment...!\n")

    for pdf_file in pdf_files:
        pdf_path = os.path.join(pdf_directory, pdf_file)
        pdf_merger.append(pdf_path)

    output_path = os.path.join(pdf_directory, folder_name+"_" + curr_time + ".pdf")

    with open(output_path, 'wb') as output_file:
        pdf_merger.write(output_file)
    print("\nPDFs merged successfully. Thank you!")

def cleanExcel():
    curr_time = datetime.now().strftime("%Y%m%d%H%M%S")

    input_files = []

    excel_directory = input("\nEnter Excel folder directory:\n")

   
    while True:
        folder_name = os.path.basename(excel_directory)
        valid_extensions = ('.xlsx', '.xls', '.csv')
        excel_files = [excel_file for excel_file in os.listdir(excel_directory) if excel_file.lower().endswith(valid_extensions)]
        excel_files.sort()

        if not excel_files:
            print("\nNo Excel or CSV files found in the directory.")
            return
        
        print("\nExcel and CSV files found in the directory:\n")
        for index, excel_file in enumerate(excel_files, start=1):
            print(f"{index}. {excel_file}")
        
        excel_choice = input("\nEnter the number of the EXCEL to clean or type 'cancel': ")
        
        if excel_choice.lower() == "cancel":
            print("\nExcel Cleaning canceled. Thank you!")
            return
        
        excel_index = int(excel_choice) - 1
        if excel_index < 0 or excel_index >= len(excel_files):
            print("Invalid Excel number.")
            return

        selected_excel = excel_files[excel_index]
        excel_path = os.path.join(excel_directory, selected_excel)

        file_name_in_color = Back.BLUE + selected_excel + Style.RESET_ALL
        print("\nCleaning Excel File: ", file_name_in_color)

        try:
            df = pd.read_excel(excel_path)
            wb = load_workbook(excel_path)
            ws = wb.active
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = cell.value
            
            output_path = os.path.join(excel_directory, f"{os.path.splitext(selected_excel)[0]}_{curr_time}.xlsx")

            df.to_excel(output_path, index=False)
            print("\nExcel file " +file_name_in_color+ " cleaning done.")
        except PermissionError:
            print("\n"+Back.RED+f"File '{selected_excel}' not valid. Skipping file..."+Style.RESET_ALL)
            continue
            
def mergeExcel():
    curr_time = datetime.now().strftime("%Y%m%d%H%M%S")

    input_files = []

    excel_directory = input("\nEnter Excel folder directory:\n")
    folder_name = os.path.basename(excel_directory)
    valid_extensions = ('.xlsx', '.xls', '.csv')
    excel_files = [excel_file for excel_file in os.listdir(excel_directory) if excel_file.lower().endswith(valid_extensions)]
    excel_files.sort()

    if not excel_files:
        print("\nNo Excel or CSV files found in the directory.")
        return

    print("\nExcel and CSV files found in the directory:\n")
    for excel_file in excel_files:
        print(excel_file)

    confirmation = input("\nDo you want to merge these Excel and CSV Files? (yes/no): ")
    if confirmation.lower() != "yes":
        print("\nExcel and CSV files merging canceled. Thank you!")
        return

    for excel_file in excel_files:
        excel_path = os.path.join(excel_directory, excel_file)
        input_files.append(excel_path)

    # Create a new output workbook
    workbook_output = openpyxl.Workbook()
    # Get the active worksheet of the output workbook
    worksheet_output = workbook_output.active

    for i, input_file in enumerate(input_files):
        # Load the input workbook into memory
        if input_file.lower().endswith(('.xlsx', '.xls')):
            try:
                workbook = openpyxl.load_workbook(input_file)
                worksheet = workbook.active
            except FileNotFoundError:
                print("\nFile '{0}' not found. Skipping file...".format(input_file))
                continue
            if i == 0:
                header_row = [cell.value for cell in worksheet[1]]
                worksheet_output.append(header_row)
            for row in worksheet.iter_rows(min_row=2):
                row_values = [cell.value for cell in row]
                worksheet_output.append(row_values)
        elif input_file.lower().endswith('.csv'):
            try:
                with open(input_file, 'r', newline='') as csvfile:
                    csvreader = csv.reader(csvfile)
                    for i, row in enumerate(csvreader):
                        if i == 0:
                            # Assume the first row in the CSV file is the header
                            header_row = row
                            worksheet_output.append(header_row)
                        else:
                            worksheet_output.append(row)
            except FileNotFoundError:
                print("\nFile '{0}' not found. Skipping file...".format(input_file))
                continue

    output_path = os.path.join(excel_directory, folder_name + "_" + curr_time + ".xlsx")
    workbook_output.save(output_path)

    print("\nMerging Excel and CSV Files has been completed. Thank you!")

def clean_and_merge():
    # Create an empty list to store the input and output file names
    file_names = []

    # Ask the user for the file names and add them to the list
    while True:
        file_name_in = input("\nEnter excel file full path without .xlsx (or press Enter to finish): ")
        if not file_name_in:
            break
        file_name_out = input("Enter output file name: ")
        file_names.append((file_name_in, file_name_out))

    # Merged output file name
    merged_output = input("Merged File name: ")

    # Get all file name in joined
    all_file_name = "\n".join(str(in_file) for in_file in file_names)

    print('\nYou have input these files: \n\n'+all_file_name)
    print("\nNew File Name: " +Back.BLUE + merged_output + Style.RESET_ALL+"")

    # Create a new output workbook
    workbook_output = openpyxl.Workbook()

    # Get the active worksheet of the output workbook
    worksheet_output = workbook_output.active

    # Loop through the list of file names, clean each file, and copy its rows to the output worksheet
    for file_name_in, file_name_out in file_names:

        file_name_in_color = Back.BLUE + file_name_in + Style.RESET_ALL
        print("\nCleaning Excel File: ", file_name_in_color)

        # Read the Excel file
        try:
            df = pd.read_excel(file_name_in+'.xlsx')
        except FileNotFoundError:
            print("\n"+Back.RED+f"File '{file_name_in}.xlsx' not found. Skipping file..."+Style.RESET_ALL)
            continue

        # Remove formatting and objects
        wb = openpyxl.load_workbook(file_name_in+'.xlsx')
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                cell.value = cell.value

        # Write the cleaned data to a new Excel file
        df.to_excel(file_name_out+'_reformated.xlsx', index=False)

        print("Excel file " +file_name_in_color+ " cleaning done.")

        # Load the input workbook into memory and select the first worksheet
        workbook = openpyxl.load_workbook(file_name_out + "_reformated.xlsx")
        worksheet = workbook.worksheets[0]

        # If this is the first input file, append its entire first row (header row) to the output worksheet
        if file_names.index((file_name_in, file_name_out)) == 0:
            header_row = [cell.value for cell in worksheet[1]]
            worksheet_output.append(header_row)

        # Append all the rows from the input worksheet (except the first row, which contains headers)
        # to the output worksheet
        for row in worksheet.iter_rows(min_row=2):
            # Get the values of each cell in the row and add them to a list
            row_values = [cell.value for cell in row]
            # Append the list of row values to the output worksheet
            worksheet_output.append(row_values)

    # Save the output workbook to the specified filename
    print("\nExcel file " +Back.BLUE + merged_output + Style.RESET_ALL+ " merging started, please wait....")
    workbook_output.save(merged_output+".xlsx")

    print("\nCleaning and merging excel files have been completed. Thank you !")

def filter_and_merge_key():
   
    curr_time = datetime.now().strftime("%Y%m%d%H%M%S")

    # Create an empty list to store the input file names
    input_files = []

    # Specify the directory containing the PDF files
    excel_directory = input("\nEnter Excel folder directory:\n")
    # Get the folder name
    folder_name = os.path.basename(excel_directory)
    # Get a list of PDF files in the directory and sort them by filename
    excel_files = [excel_file for excel_file in os.listdir(excel_directory) if excel_file.lower().endswith('.xlsx')]
    excel_files.sort()

    if not excel_files:
        print("\nNo Excel files found in the directory.")
        return

    print("\nExcel files found in the directory:\n")
    for excel_file in excel_files:
        print(excel_file)

    # Set the name of the column to filter and the keywords to look for
    column_to_filter = input("\nEnter filter column name: ")
    keywords_to_filter = input("Enter " + column_to_filter + " filter keywords (separated by commas): ").split(",")
    keywords_to_filter = [keyword.strip().lower() for keyword in keywords_to_filter]

    confirmation = input("\nDo you want to filter and merge these Excel Files? (yes/no): ")
    if confirmation.lower() != "yes":
        print("\nExcel files filtering and merging canceled. Thank you!")
        return
    
    for excel_file in excel_files:
        excel_path = os.path.join(excel_directory, excel_file)
        input_files.append(excel_path)

    # Initialize a list to store the filtered dataframes
    dfs = []

    # Loop through the Excel files and filter the desired column by the keywords
    for file in input_files:
        try:
            df = pd.read_excel(file, sheet_name=0)
        except FileNotFoundError:
            print("\n" + Back.RED + f"File '{file}.xlsx' not found. Skipping file..." + Style.RESET_ALL)
            continue
        df[column_to_filter] = df[column_to_filter].astype(str).str.lower()
        if not keywords_to_filter:
            df_filtered = df[df[column_to_filter]==""]
        else:
            df_filtered = df[df[column_to_filter].isin(keywords_to_filter)]

        dfs.append(df_filtered)

    # Concatenate the filtered dataframes into a single dataframe
    merged_df = pd.concat(dfs, axis=0)

    # Save the merged dataframe to a new Excel file
    output_file_name = os.path.join(excel_directory, folder_name+"_" + "_".join(keywords_to_filter) + "_" + curr_time + ".xlsx")
    with pd.ExcelWriter(output_file_name) as writer:
        merged_df.to_excel(writer, index=False, sheet_name="Filtered")

    print('\nFiltered in ' + Back.BLUE + column_to_filter + Style.RESET_ALL + ' (column) by ' +
          Back.BLUE + ", ".join(keywords_to_filter) + Style.RESET_ALL + ' (column keywords) and excel file merged done. Thank you !') 

def merge_sheets():
    curr_time = datetime.now().strftime("%Y%m%d%H%M%S")
    # Create an empty list to store the input file names
    excel_file_names = []

    # Ask the user to enter the input file names one by one, until they enter 'done'
    while True:
        filename = input("Enter excel file full path without .xlsx (or press Enter to finish): ")
        if not filename:
            break
        excel_file_names.append(filename)

    # Get all file name in joined
    all_file_name = "\n".join(str(in_file) for in_file in excel_file_names)

    print('\nYou have input these files: \n\n' + all_file_name)
    print("\nSheets merging started, please wait....")

    # Create a new sheet to merge all the sheets into.
    merged_sheet = pd.DataFrame()

    # Iterate over all the Excel files and merge their sheets into the new sheet.
    for excel_file in excel_file_names:
        sheets = pd.ExcelFile(excel_file+".xlsx").sheet_names
        for sheet in sheets:
            merged_sheet = pd.concat([merged_sheet, pd.read_excel(excel_file+".xlsx", sheet_name=sheet)], ignore_index=True)

        # Save each merged sheet to a new Excel file
        merged_sheet.to_excel(f"{excel_file}_{curr_time}.xlsx", sheet_name="merged_sheet")
        merged_sheet = pd.DataFrame()  # Reset the merged_sheet DataFrame for the next file

    print('\nSheets of excel file '+Back.BLUE + ", ".join(excel_file_names) + Style.RESET_ALL+' merged done. Thank you !') 

# Call the initial function

try:
    initial()
except KeyboardInterrupt:
    print("\n\n" + Back.RED + "Keyboard interrupted ! Exiting by the user. Thank you !" + Style.RESET_ALL + "\n")
    exit(1)

